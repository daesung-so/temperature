from flask import Flask, request, jsonify, send_file, make_response
from flask_cors import CORS
import openpyxl
from datetime import datetime, time as time_cls
from copy import copy
import io
import os
import json
import re
import zipfile
import traceback

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}}, supports_credentials=False)

@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin']  = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    return response


FLOOR_SCHEMA = {
    '2층': { 'units': [0,1,2,3,4,5,6,7], 'start_col': 9 },
    '3층': { 'units': [0,1,2,3],         'start_col': 9 },
    '5층': { 'units': [0,1,2,3],         'start_col': 9 },
    '6층': { 'units': [0,1,2,4,5,6,7],   'start_col': 9 },
    '8층': { 'units': [0,1,2,3],         'start_col': 9 },
}

DATA_SHEETS = {'2층', '3층', '5층', '6층', '8층'}


def restore_drawings(processed_buf, original_bytes):
    """openpyxl이 변경한 파일들을 원본에서 최대한 복원
    - 데이터 입력 시트(2층, 3층, 5층, 6층, 8층): 처리본 사용 (새 행 포함)
    - 그 외 모든 파일: 원본 그대로 복원 (workbook, styles, sharedStrings, drawings, sheet1, sheet7 등)
    """
    import zipfile, re
    
    processed_buf.seek(0)

    # 데이터 시트 식별 (workbook.xml과 rels에서 매핑)
    DATA_SHEETS = {'2층', '3층', '5층', '6층', '8층'}
    data_sheet_paths = set()  # 처리본에서 가져올 sheet xml 경로

    with zipfile.ZipFile(io.BytesIO(original_bytes), 'r') as orig_zip:
        wb_xml = orig_zip.read('xl/workbook.xml').decode('utf-8')
        wb_rels = orig_zip.read('xl/_rels/workbook.xml.rels').decode('utf-8')

        sheet_rid = {}
        for m in re.finditer(r'<sheet\s+([^>]+?)/?>', wb_xml):
            attrs = m.group(1)
            n = re.search(r'name="([^"]+)"', attrs)
            r_ = re.search(r'r:id="(rId\d+)"', attrs)
            if n and r_:
                sheet_rid[n.group(1)] = r_.group(1)

        rid_target = {}
        for m in re.finditer(r'<Relationship\s+([^>]+?)/>', wb_rels):
            attrs = m.group(1)
            id_m = re.search(r'Id="(rId\d+)"', attrs)
            tgt_m = re.search(r'Target="(worksheets/sheet\d+\.xml)"', attrs)
            if id_m and tgt_m:
                rid_target[id_m.group(1)] = tgt_m.group(1)

        for sheet_name in DATA_SHEETS:
            if sheet_name in sheet_rid:
                rid = sheet_rid[sheet_name]
                if rid in rid_target:
                    data_sheet_paths.add('xl/' + rid_target[rid])

    # 새 zip 작성: 원본을 베이스로, 데이터 시트만 처리본에서 가져오기
    new_buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(original_bytes), 'r') as orig_zip:
        with zipfile.ZipFile(processed_buf, 'r') as proc_zip:
            with zipfile.ZipFile(new_buf, 'w', zipfile.ZIP_DEFLATED) as new_zip:
                proc_files = set(proc_zip.namelist())
                
                # 원본의 모든 파일을 베이스로 사용
                for name in orig_zip.namelist():
                    if name in data_sheet_paths and name in proc_files:
                        # 데이터 시트는 처리본 사용
                        new_zip.writestr(name, proc_zip.read(name))
                    else:
                        # 그 외는 원본 그대로
                        new_zip.writestr(name, orig_zip.read(name))

    new_buf.seek(0)
    return new_buf


@app.route('/', methods=['GET'])
def index():
    return jsonify({'service': 'temperature-server', 'status': 'ok'})


@app.route('/health', methods=['GET'])
def health():
    return jsonify({'ok': True})


def make_summary_formulas(schema, row_num):
    """각 층의 요약 수식 생성 (이전 행에 수식이 없어도 동작)"""
    from openpyxl.utils import get_column_letter
    units = schema['units']
    sc = schema['start_col']  # 9
    
    # 온도/습도 셀 참조
    temp_refs = [get_column_letter(sc + i*2)     + str(row_num) for i in range(len(units))]
    hum_refs  = [get_column_letter(sc + i*2 + 1) + str(row_num) for i in range(len(units))]
    
    # 요약 컬럼 시작
    s = sc + len(units) * 2  # 2층=25(Y), 3층=17(Q)
    
    # 호기 위치 표시용: 각 호기의 라벨이 있는 셀 ($I$2, $K$2, ...)
    pos_refs_t = [f'$' + get_column_letter(sc + i*2) + '$2' for i in range(len(units))]
    pos_refs_h = [f'$' + get_column_letter(sc + i*2 + 1) + '$2' for i in range(len(units))]
    
    Y  = get_column_letter(s)       # 평균온도
    Z  = get_column_letter(s + 1)   # 평균습도
    AA = get_column_letter(s + 2)   # 최고온도
    AB = get_column_letter(s + 3)   # 최저습도
    
    # 판정: IF(OR(MIN<21, MAX>27), "불량", "양호")  - 온도 기준
    # 판정: IF(OR(MIN<30, MAX>70), "불량", "양호")  - 습도 기준
    judge_t = f'=IF(OR(MIN({",".join(temp_refs)})<21,MAX({",".join(temp_refs)})>27),"불량","양호")'
    judge_h = f'=IF(OR(MIN({",".join(hum_refs)})<30,MAX({",".join(hum_refs)})>70),"불량","양호")'
    
    # 위치: IF(MAX=I, $I$2, IF(MAX=K, $K$2, ...))
    def make_pos_formula(max_ref, val_refs, pos_refs):
        # IF 중첩
        result = f'""'
        for vr, pr in zip(reversed(val_refs), reversed(pos_refs)):
            result = f'IF({max_ref}={vr},{pr},{result})'
        return '=' + result
    
    pos_t = make_pos_formula(f'{AA}{row_num}', temp_refs, pos_refs_t)
    pos_h = make_pos_formula(f'{AB}{row_num}', hum_refs, pos_refs_h)
    
    return {
        s:     f'=AVERAGE({",".join(temp_refs)})',  # Y: 평균온도
        s + 1: f'=AVERAGE({",".join(hum_refs)})',   # Z: 평균습도
        s + 2: f'=MAX({",".join(temp_refs)})',      # AA: 최고온도
        s + 3: f'=MIN({",".join(hum_refs)})',       # AB: 최저습도
        s + 4: judge_t,                             # AC: 온도판정
        s + 5: judge_h,                             # AD: 습도판정
        s + 7: pos_t,                               # AF: 온도위치
        s + 8: pos_h,                               # AG: 습도위치
    }


@app.route('/update-xlsm', methods=['POST', 'OPTIONS'])
def update_xlsm():
    if request.method == 'OPTIONS':
        return make_response('', 204)

    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        if 'data' not in request.form:
            return jsonify({'error': 'No data provided'}), 400

        xlsm_file = request.files['file']
        record = json.loads(request.form['data'])

        original_bytes = xlsm_file.read()
        buf = io.BytesIO(original_bytes)
        wb = openpyxl.load_workbook(
            buf, keep_vba=True, keep_links=False, data_only=False
        )

        date_str  = record.get('date', '')
        day_str   = record.get('day', '')
        time_val  = record.get('timeVal', '00')
        shift     = record.get('shift', '주간')
        inspector = record.get('inspector', '')
        floors_data = record.get('floors', {})

        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            mon = str(date_obj.month)
            day = str(date_obj.day).zfill(2)
        except:
            return jsonify({'error': f'Invalid date: {date_str}'}), 400

        key = mon + day + day_str + shift
        hh = int(time_val)
        time_obj = time_cls(hh, 0)

        for floor_name, schema in FLOOR_SCHEMA.items():
            if floor_name not in wb.sheetnames:
                continue
            ws = wb[floor_name]
            floor_data = floors_data.get(floor_name, {})

            # 마지막 NO/행 찾기
            last_no = 0
            last_row = ws.max_row
            for r in range(ws.max_row, 3, -1):
                val = ws.cell(row=r, column=2).value
                if isinstance(val, (int, float)):
                    last_no = int(val)
                    last_row = r
                    break
            new_row = last_row + 1
            new_no  = last_no + 1

            # 수식 있는 템플릿 행 찾기
            summary_col = schema['start_col'] + len(schema['units']) * 2
            template_row = last_row
            for r in range(last_row, 3, -1):
                val = ws.cell(row=r, column=summary_col).value
                if isinstance(val, str) and val.startswith('='):
                    template_row = r
                    break

            # 기본 정보
            ws.cell(row=new_row, column=1).value = key
            ws.cell(row=new_row, column=2).value = new_no
            ws.cell(row=new_row, column=3).value = mon
            ws.cell(row=new_row, column=4).value = day
            ws.cell(row=new_row, column=5).value = day_str
            ws.cell(row=new_row, column=6).value = shift
            ws.cell(row=new_row, column=7).value = time_obj
            ws.cell(row=new_row, column=7).number_format = 'h:mm'
            ws.cell(row=new_row, column=8).value = inspector

            # 호기별 온도/습도
            units = schema['units']
            sc = schema['start_col']
            for i, u in enumerate(units):
                d = floor_data.get(str(u), {})
                t_val = d.get('temp', '')
                h_val = d.get('hum', '')
                t = float(t_val) if t_val != '' else None
                h = float(h_val) if h_val != '' else None
                ws.cell(row=new_row, column=sc + i*2).value     = t
                ws.cell(row=new_row, column=sc + i*2 + 1).value = h
            
            # 요약 수식 직접 생성 (이전 행 의존 X)
            summary_formulas = make_summary_formulas(schema, new_row)
            for col, formula in summary_formulas.items():
                ws.cell(row=new_row, column=col).value = formula

            # 스타일 + 수식 복사 (template_row 기준)
            data_end_col = schema['start_col'] + len(schema['units']) * 2 - 1
            old_r = str(template_row)
            new_r = str(new_row)

            for col in range(1, ws.max_column + 1):
                src = ws.cell(row=template_row, column=col)
                dst = ws.cell(row=new_row, column=col)

                # 스타일
                if src.has_style:
                    dst._style = copy(src._style)
                    dst.number_format = src.number_format

                # 수식 복사 (요약 컬럼 이후)
                if col > data_end_col and isinstance(src.value, str) and src.value.startswith('='):
                    new_formula = re.sub(
                        r'([A-Z]+)' + old_r + r'\b',
                        lambda m: m.group(1) + new_r,
                        src.value
                    )
                    dst.value = new_formula

        # 파일명 (4자리 연도)
        date_part = date_str.replace('-', '')
        filename = f'일일점검사항_v8__{date_part}_{time_val}시.xlsm'

        out_buf = io.BytesIO()
        wb.save(out_buf)
        out_buf.seek(0)

        # ZIP 조작으로 보조 파일 복원
        out_buf = restore_drawings(out_buf, original_bytes)

        return send_file(
            out_buf,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.ms-excel.sheet.macroEnabled.12'
        )

    except Exception as e:
        print('ERROR:', str(e))
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
