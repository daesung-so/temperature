from __future__ import annotations

import io
import re
import tempfile
from copy import copy
from datetime import datetime, time
from pathlib import Path
from typing import Any
from zipfile import ZipFile

import requests
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

API_URL = "https://so-schedule.sungsu.workers.dev/api/temperature"

FLOORS: dict[str, dict[str, list[int]]] = {
    "2층": {"units": [0, 1, 2, 3, 4, 5, 6, 7]},
    "3층": {"units": [0, 1, 2, 3]},
    "5층": {"units": [0, 1, 2, 3]},
    "6층": {"units": [0, 1, 2, 4, 5, 6, 7]},
    "8층": {"units": [0, 1, 2, 3]},
}

LIMITS = {"tMin": 21, "tMax": 27, "hMin": 30, "hMax": 70}

app = FastAPI(title="Sungsu XLSM Processor", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.post("/process-xlsm")
def process_xlsm(
    file: UploadFile = File(...),
    record_mode: str = Form("latest"),
    record_date: str | None = Form(None),
    record_time_val: str | None = Form(None),
) -> StreamingResponse:
    suffix = Path(file.filename or "upload.xlsm").suffix.lower()
    if suffix != ".xlsm":
        raise HTTPException(status_code=400, detail=".xlsm 파일만 업로드할 수 있습니다.")

    raw = file.file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="빈 파일입니다.")
    if b"xl/vbaProject.bin" not in raw:
        raise HTTPException(status_code=400, detail="매크로(vbaProject.bin)가 없는 XLSM 파일입니다.")

    record = fetch_record(record_mode=record_mode, record_date=record_date, record_time_val=record_time_val)

    try:
        output_bytes, summary = inject_record(raw, record, file.filename or "upload.xlsm")
    except HTTPException:
        raise
    except Exception as exc:  # pragma: no cover - runtime guard
        raise HTTPException(status_code=500, detail=f"XLSM 처리 실패: {exc}") from exc

    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{summary['download_name']}",
        "X-Process-Summary": summary["message"],
    }
    return StreamingResponse(
        io.BytesIO(output_bytes),
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
        headers=headers,
    )


@app.get("/latest-record")
def latest_record() -> JSONResponse:
    record = fetch_record(record_mode="latest", record_date=None, record_time_val=None)
    return JSONResponse(record)


def fetch_record(record_mode: str, record_date: str | None, record_time_val: str | None) -> dict[str, Any]:
    try:
        resp = requests.get(API_URL, timeout=15)
        resp.raise_for_status()
    except requests.RequestException as exc:
        raise HTTPException(status_code=502, detail=f"저장 데이터 조회 실패: {exc}") from exc

    payload = resp.json()
    records = payload if isinstance(payload, list) else payload.get("record", [])
    if not records:
        raise HTTPException(status_code=404, detail="저장된 점검 데이터가 없습니다.")

    if record_mode == "explicit":
        if not record_date or not record_time_val:
            raise HTTPException(status_code=400, detail="explicit 모드에는 record_date와 record_time_val이 필요합니다.")
        for record in reversed(records):
            if record.get("date") == record_date and str(record.get("timeVal")) == str(record_time_val):
                return record
        raise HTTPException(status_code=404, detail="선택한 날짜/시간 기록을 찾지 못했습니다.")

    return records[-1]


def inject_record(raw: bytes, record: dict[str, Any], original_name: str) -> tuple[bytes, dict[str, str]]:
    with tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False) as src:
        src.write(raw)
        src_path = Path(src.name)

    wb = load_workbook(src_path, keep_vba=True)

    touched: list[str] = []
    row_messages: list[str] = []
    for floor in FLOORS:
        if floor not in wb.sheetnames:
            continue
        ws = wb[floor]
        row_no, action = upsert_floor_sheet(ws, floor, record)
        touched.append(floor)
        row_messages.append(f"{floor} {row_no}행 {action}")

    with tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False) as out:
        out_path = Path(out.name)
    wb.save(out_path)

    result_bytes = out_path.read_bytes()
    if b"xl/vbaProject.bin" not in result_bytes:
        raise HTTPException(status_code=500, detail="처리 후 매크로가 보존되지 않았습니다.")

    download_name = build_download_name(original_name, record)
    summary = {
        "download_name": download_name,
        "message": f"반영 완료: {', '.join(row_messages)}",
    }
    return result_bytes, summary


def upsert_floor_sheet(ws, floor: str, record: dict[str, Any]) -> tuple[int, str]:
    target_row = find_existing_row(ws, record)
    action = "덮어쓰기" if target_row else "추가"
    if target_row is None:
        template_row = find_last_data_row(ws)
        target_row = template_row + 1
        ws.insert_rows(target_row, amount=1)
        copy_row_style_and_formulas(ws, template_row, target_row)
        # NO / 구분명은 새 행 번호 기준으로 갱신
        set_cell_value(ws, target_row, 2, target_row - 3)
        set_cell_value(ws, target_row, 1, make_label(record, ws, target_row))
    else:
        # 같은 날짜/시간 덮어쓰기여도 수식 칼럼은 유지되도록 기존 수식은 건드리지 않음
        pass

    apply_record_to_row(ws, floor, target_row, record)
    refresh_row_formulas(ws, floor, target_row)
    return target_row, action


def find_last_data_row(ws) -> int:
    row = ws.max_row
    while row >= 4:
        if ws.cell(row, 2).value not in (None, ""):
            return row
        row -= 1
    return 4


def find_existing_row(ws, record: dict[str, Any]) -> int | None:
    month = int(record["date"][5:7])
    day = int(record["date"][8:10])
    shift = record.get("shift", "")
    time_val = str(record.get("timeVal", ""))

    for row in range(4, ws.max_row + 1):
        r_month = normalize_int(ws.cell(row, 3).value)
        r_day = normalize_int(ws.cell(row, 4).value)
        r_shift = str(ws.cell(row, 6).value or "").strip()
        cell_time = ws.cell(row, 7).value
        r_time_val = normalize_hour(cell_time)
        if r_month == month and r_day == day and r_shift == shift and r_time_val == time_val:
            return row
    return None


def normalize_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, int):
        return value
    try:
        return int(str(value).strip())
    except ValueError:
        return None


def normalize_hour(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, time):
        return f"{value.hour:02d}"
    text = str(value).strip()
    match = re.search(r"(\d{1,2})", text)
    return f"{int(match.group(1)):02d}" if match else text


def copy_row_style_and_formulas(ws, src_row: int, dst_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(src_row, col)
        dst_cell = ws.cell(dst_row, col)

        if src_cell.has_style:
            dst_cell._style = copy(src_cell._style)
        if src_cell.number_format:
            dst_cell.number_format = copy(src_cell.number_format)
        if src_cell.font:
            dst_cell.font = copy(src_cell.font)
        if src_cell.fill:
            dst_cell.fill = copy(src_cell.fill)
        if src_cell.border:
            dst_cell.border = copy(src_cell.border)
        if src_cell.alignment:
            dst_cell.alignment = copy(src_cell.alignment)
        if src_cell.protection:
            dst_cell.protection = copy(src_cell.protection)

        if isinstance(src_cell.value, str) and src_cell.value.startswith("="):
            dst_cell.value = shift_formula_rows(src_cell.value, src_row, dst_row)
        else:
            dst_cell.value = src_cell.value

    if src_row in ws.row_dimensions:
        src_dim = ws.row_dimensions[src_row]
        dst_dim = ws.row_dimensions[dst_row]
        dst_dim.height = src_dim.height
        dst_dim.hidden = src_dim.hidden


CELL_REF_RE = re.compile(r"(?<![A-Z])((\$?[A-Z]{1,3})(\$?)(\d+))")


def shift_formula_rows(formula: str, src_row: int, dst_row: int) -> str:
    delta = dst_row - src_row

    def repl(match: re.Match[str]) -> str:
        full_ref = match.group(1)
        col = match.group(2)
        row_lock = match.group(3)
        row_num = int(match.group(4))
        if row_lock == "$":
            return full_ref
        return f"{col}{row_num + delta}"

    return CELL_REF_RE.sub(repl, formula)


def apply_record_to_row(ws, floor: str, row: int, record: dict[str, Any]) -> None:
    date_obj = datetime.strptime(record["date"], "%Y-%m-%d")
    day_name = record.get("day") or ["월", "화", "수", "목", "금", "토", "일"][date_obj.weekday()]
    shift = record.get("shift", "")
    time_val = str(record.get("timeVal", "00"))
    inspector = record.get("inspector") or "(미입력)"

    set_cell_value(ws, row, 1, make_label(record, ws, row))
    set_cell_value(ws, row, 2, row - 3)
    set_cell_value(ws, row, 3, str(date_obj.month))
    set_cell_value(ws, row, 4, f"{date_obj.day:02d}")
    set_cell_value(ws, row, 5, day_name)
    set_cell_value(ws, row, 6, shift)
    set_cell_value(ws, row, 7, time(int(time_val), 0))
    set_cell_value(ws, row, 8, inspector)

    unit_values = record.get("floors", {}).get(floor, {})
    units = FLOORS[floor]["units"]
    start_col = 9
    for idx, unit in enumerate(units):
        base_col = start_col + idx * 2
        datum = unit_values.get(str(unit), unit_values.get(unit, {"temp": "", "hum": ""}))
        temp_val = parse_numeric_or_text(datum.get("temp", ""))
        hum_val = parse_numeric_or_text(datum.get("hum", ""))
        set_cell_value(ws, row, base_col, temp_val)
        set_cell_value(ws, row, base_col + 1, hum_val)
        apply_limit_fill(ws.cell(row, base_col), temp_val, LIMITS["tMin"], LIMITS["tMax"])
        apply_limit_fill(ws.cell(row, base_col + 1), hum_val, LIMITS["hMin"], LIMITS["hMax"])


def set_cell_value(ws, row: int, col: int, value: Any) -> None:
    ws.cell(row, col).value = value


RED_FILL = PatternFill(fill_type="solid", fgColor="FFF1F2")
NO_FILL = PatternFill(fill_type=None)


def apply_limit_fill(cell, value: Any, low: float, high: float) -> None:
    if isinstance(value, (int, float)):
        cell.fill = copy(RED_FILL if (value < low or value > high) else NO_FILL)


def parse_numeric_or_text(value: Any) -> Any:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    try:
        return float(text)
    except ValueError:
        return text


def make_label(record: dict[str, Any], ws, row: int) -> str:
    month = int(record["date"][5:7])
    day = int(record["date"][8:10])
    day_name = record.get("day") or ""
    shift = "주간" if str(record.get("timeVal")) != "22" else "야간"
    return f"{month:02d}{day_name}{shift}"


def refresh_row_formulas(ws, floor: str, row: int) -> None:
    # 수식 칼럼은 바로 위 행 기준 복사된 상태이므로, 덮어쓰기일 때만 수식 재생성
    if row <= 4:
        return
    formula_start = first_formula_col(ws)
    if formula_start is None:
        return
    prev_row = row - 1
    for col in range(formula_start, ws.max_column + 1):
        src_val = ws.cell(prev_row, col).value
        dst_cell = ws.cell(row, col)
        if isinstance(src_val, str) and src_val.startswith("="):
            dst_cell.value = shift_formula_rows(src_val, prev_row, row)


def first_formula_col(ws) -> int | None:
    for col in range(1, ws.max_column + 1):
        if isinstance(ws.cell(4, col).value, str) and ws.cell(4, col).value.startswith("="):
            return col
    return None


def build_download_name(original_name: str, record: dict[str, Any]) -> str:
    base = Path(original_name).stem
    date_token = record["date"].replace("-", "")
    time_token = f"{str(record.get('timeVal', '00')).zfill(2)}시"
    replacement = f"__{date_token}_{time_token}"
    if re.search(r"__\d{8}_\d{2}시$", base):
        new_base = re.sub(r"__\d{8}_\d{2}시$", replacement, base)
    else:
        new_base = f"{base}{replacement}"
    return f"{new_base}.xlsm"
